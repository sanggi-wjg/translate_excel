from openpyxl import load_workbook

from apps.utils import ExcelReadError


class ReadExcel:

    def read(self, path: str) -> list:
        read_result = []
        workbook = load_workbook(filename = path, data_only = True)

        try:
            sheet = workbook.worksheets[0]

            for col in sheet.iter_cols():
                for cell in col:
                    if cell.value is not None and cell.value not in ['!', '@', '#', '$', '%', '^', '&', '*', '(', ')', '-', '_', '+', '=']:
                        read_result.append({
                            'column': cell.column_letter,
                            'row'   : str(cell.row),
                            'value' : str(cell.value)
                        })

        except Exception:
            raise ExcelReadError

        finally:
            workbook.close()

        return read_result
